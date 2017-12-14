import matplotlib.pyplot as plt
import numpy as np
from sklearn.linear_model import LinearRegression
from openpyxl import load_workbook

wb = load_workbook("template.xlsx")
sheet = wb.get_sheet_by_name("系統百分比結果")

temperatures=[]
iced_tea_sales=[]


for row in sheet.iter_rows(min_col=11,min_row=2, max_col=11, max_row=1614):
    for cell in row:
        #print(cell.value)
        temperatures.append(cell.value)
print("===========================")
for row in sheet.iter_rows(min_col=13 ,min_row=2, max_col=13, max_row=1614):
    for cell in row:
        #print(cell.value)
        iced_tea_sales.append(cell.value)

#temperatures = np.array([29, 28, 34, 31, 25, 29, 32, 31, 24, 33, 25, 31, 26, 30])
#iced_tea_sales = np.array([77, 62, 93, 84, 59, 64, 80, 75, 58, 91, 51, 73, 65, 84])

lm = LinearRegression()
#lm.fit(np.reshape(temperatures, (len(temperatures), 1)), np.reshape(iced_tea_sales, (len(iced_tea_sales), 1)))
lm.fit(np.reshape(iced_tea_sales, (len(iced_tea_sales), 1)),np.reshape(temperatures, (len(temperatures), 1)))



# 新的氣溫
to_be_predicted = np.array([60])
predicted = lm.predict(np.reshape(to_be_predicted, (len(to_be_predicted), 1)))
print(predicted)
#to_be_predicted = np.array([511.9759])
#predicted_sales = lm.predict(np.reshape(to_be_predicted, (len(to_be_predicted), 1)))
#print(predicted_sales)



# 視覺化
#plt.scatter(temperatures, iced_tea_sales, color='black')
#plt.plot(temperatures, lm.predict(np.reshape(temperatures, (len(temperatures), 1))), color='blue', linewidth=3)
plt.scatter(temperatures,iced_tea_sales , color='black')
plt.plot(lm.predict(np.reshape(iced_tea_sales, (len(iced_tea_sales), 1))),iced_tea_sales , color='blue', linewidth=3)

plt.plot(predicted,to_be_predicted , color = 'red', marker = '*', markersize = 20)
#plt.xticks(())
#plt.yticks(())
plt.ylim((0, 350))
plt.xlabel('RT')
plt.ylabel('hen')
plt.show()
